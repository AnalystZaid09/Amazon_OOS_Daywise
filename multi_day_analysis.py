import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime

# Page configuration
st.set_page_config(page_title="Multi-Day Business Report Tool", layout="wide")

st.markdown("""
<style>
    .main {
        background-color: #f5f7f9;
    }
    .stButton>button {
        width: 100%;
        border-radius: 5px;
        height: 3em;
        background-color: #007bff;
        color: white;
    }
</style>
""", unsafe_allow_html=True)

st.title("📊 Multi-Day Business Report Tool")
st.markdown("Consolidate business reports for 7, 15, 30, 45, 60, 75, and 90-day intervals.")

# Sidebar for file uploads
with st.sidebar:
    st.header("📁 Upload Reports")
    
    intervals = [7, 15, 30, 45, 60, 75, 90]
    br_files = {}
    for day in intervals:
        br_files[day] = st.file_uploader(f"Business Report - {day} Days", type=["csv", "xlsx"], key=f"br_{day}")
    
    st.divider()
    pm_file = st.file_uploader("Purchase Master (Excel)", type=["xlsx", "xls"])
    inv_file = st.file_uploader("Inventory Report (CSV)", type=["csv"])
    res_file = st.file_uploader("Reserved Inventory (CSV)", type=["csv"], key="res_file")
    list_file = st.file_uploader("Listing Report (CSV/Excel)", type=["csv", "xlsx"], key="list_file")
    
    st.divider()
    st.markdown("### 📊 DOC Color Legend")
    st.markdown("""
    🔴 **Red (0-7)**: Critical  
    🟠 **Orange (7-15)**: Low  
    🟢 **Green (15-30)**: Optimal  
    🟡 **Yellow (30-45)**: Monitor  
    🔵 **Sky Blue (45-60)**: High  
    🟤 **Brown (60-90)**: Excess  
    ⬛ **Black (>90)**: Overstocked
    """)
    
    process_btn = st.button("🚀 Process Multi-Day Reports")

def clean_numeric_col(df, col):
    if col in df.columns:
        series = df[col]
        if pd.api.types.is_numeric_dtype(series):
            return series.fillna(0)
        
        # If not numeric, clean it
        cleaned = (
            series.astype(str)
            .str.replace(r"[₹, %]", "", regex=True)
            .str.strip()
        )
        # Handle empty strings after replace
        cleaned = cleaned.replace('', '0')
        return pd.to_numeric(cleaned, errors="coerce").fillna(0)
    return pd.Series(0, index=df.index)

def apply_doc_styling(df):
    """
    Apply background color to DOC columns for Streamlit display.
    """
    def style_doc(val):
        try:
            val = float(val)
            if val <= 7:
                return "background-color: #FF0000; color: white" # Red
            elif val <= 15:
                return "background-color: #FFA500; color: black" # Orange
            elif val <= 30:
                return "background-color: #008000; color: white" # Green
            elif val <= 45:
                return "background-color: #FFFF00; color: black" # Yellow
            elif val <= 60:
                return "background-color: #87CEEB; color: black" # Sky Blue
            elif val <= 90:
                return "background-color: #A52A2A; color: white" # Brown
            else:
                return "background-color: #000000; color: white" # Black
        except:
            return ""

    doc_cols = [c for c in df.columns if "DOC" in c]
    if not doc_cols:
        return df
        
    return df.style.map(style_doc, subset=doc_cols)

def create_stock_pivot(df):
    if df.empty: return pd.DataFrame()
    df_copy = df.copy()
    
    # Map Max columns to base names for pivot if they exist
    rename_map = {}
    if "DOC (Max)" in df_copy.columns: 
        rename_map["DOC (Max)"] = "DOC"
    elif "DOC" in df_copy.columns:
        pass # Already named correctly
    else:
        # Check if there are any other DOC columns we can use as a fallback
        any_doc = next((c for c in df_copy.columns if "DOC" in c), None)
        if any_doc: rename_map[any_doc] = "DOC"

    if "Max DRR" in df_copy.columns: 
        rename_map["Max DRR"] = "DRR"
    elif "DRR" in df_copy.columns:
        pass
    else:
        any_drr = next((c for c in df_copy.columns if "DRR" in c), None)
        if any_drr: rename_map[any_drr] = "DRR"

    if rename_map:
        df_copy = df_copy.rename(columns=rename_map)
        
    # Identify which of our target values actually exist now
    target_values = ["DOC", "DRR", "CP"]
    available_values = [v for v in target_values if v in df_copy.columns]
    
    if not available_values:
        return pd.DataFrame()

    for col in available_values:
        df_copy[col] = pd.to_numeric(df_copy[col], errors="coerce").fillna(0)
    
    # Ensure index columns exist
    index_cols = [c for c in ["Brand", "(Parent) ASIN"] if c in df_copy.columns]
    if not index_cols:
        # Fallback to SKU if (Parent) ASIN is missing
        index_cols = [c for c in ["Brand", "SKU"] if c in df_copy.columns]
    
    if not index_cols:
        return pd.DataFrame()

    pivot = pd.pivot_table(
        df_copy,
        index=index_cols,
        values=available_values,
        aggfunc="sum",
        margins=True,
        margins_name="Grand Total"
    ).reset_index()
    
    # Precise rename for final display
    pivot_rename = {v: f"Sum of {v}" for v in available_values}
    pivot.rename(columns=pivot_rename, inplace=True)
    return pivot

def process_br(file, days):
    if file.name.endswith(".csv"):
        df = pd.read_csv(file)
    else:
        df = pd.read_excel(file)
    
    # Cleaning
    # cl.py uses Total Order Items + B2B as per snippet, but also has Units Ordered
    df["Units Ordered"] = clean_numeric_col(df, "Units Ordered")
    df["Units Ordered - B2B"] = clean_numeric_col(df, "Units Ordered - B2B")
    df["Total Order Items"] = clean_numeric_col(df, "Total Order Items")
    df["Total Order Items - B2B"] = clean_numeric_col(df, "Total Order Items - B2B")
    df["Page Views - Total"] = clean_numeric_col(df, "Page Views - Total")
    df["Page Views - Total - B2B"] = clean_numeric_col(df, "Page Views - Total - B2B")
    df["Sessions - Total"] = clean_numeric_col(df, "Sessions - Total")
    df["Buy Box Percentage"] = clean_numeric_col(df, "Buy Box Percentage")
    df["Unit Session Percentage"] = clean_numeric_col(df, "Unit Session Percentage")
    
    # cl.py logic for Sales Qty
    df["Period Sales Qty"] = df["Total Order Items"] + df["Total Order Items - B2B"]
    if df["Period Sales Qty"].sum() == 0:
        # Fallback to Units Ordered if Total Order Items is missing
        df["Period Sales Qty"] = df["Units Ordered"] + df["Units Ordered - B2B"]

    # Find SKU column in BR
    possible_sku_cols = ["SKU", "sku", "Seller SKU"]
    sku_col_br = next((c for c in possible_sku_cols if c in df.columns), None)
    
    if not sku_col_br:
        st.error(f"Could not find a SKU column in Business Report for {days} days.")
        st.stop()
    
        # Normalize SKU for potential reporting, but we will group by ASIN
    df[sku_col_br] = df[sku_col_br].astype(str).str.strip().str.upper()

    # Find ASIN in BR
    asin_col_br = next((c for c in ["(Parent) ASIN", "ASIN", "asin"] if c in df.columns), None)
    if not asin_col_br:
        st.error(f"Could not find an ASIN column in Business Report for {days} days.")
        st.stop()
    
    # Normalize ASIN
    df[asin_col_br] = df[asin_col_br].astype(str).str.strip().str.upper()

    # Total Page Views (B2C + B2B)
    df["Period Page Views"] = df["Page Views - Total"] + df["Page Views - Total - B2B"]

    # Pivot
    agg_dict = {
        "Period Sales Qty": "sum",
        "Period Page Views": "sum",
        "Sessions - Total": "sum",
        "Buy Box Percentage": "mean",
        "Unit Session Percentage": "mean"
    }

    # Group by ASIN
    pivot = df.groupby(asin_col_br).agg(agg_dict).reset_index()
    
    # Rename columns with days prefix
    cols = [
        "(Parent) ASIN", 
        f"{days} Days Sales Qty", 
        f"{days} Days Page Views",
        f"{days} Days Sessions",
        f"{days} Days Buy Box %",
        f"{days} Days Unit Session %"
    ]

    pivot.columns = cols
    return pivot

def create_excel(df_dict, intervals):
    """
    df_dict: Dictionary of {SheetName: DataFrame}
    """
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in df_dict.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Ranges and Colors (cl.py 7-color logic)
            colors = {
                "critical": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"), # Red 0-7
                "low": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),      # Orange 7-15
                "optimal": PatternFill(start_color="008000", end_color="008000", fill_type="solid"),  # Green 15-30
                "monitor": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),  # Yellow 30-45
                "high": PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),     # Sky Blue 45-60
                "excess": PatternFill(start_color="A52A2A", end_color="A52A2A", fill_type="solid"),   # Brown 60-90
                "overstock": PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Black >90
            }
            font_white = Font(color="FFFFFF", bold=True)
            font_black = Font(color="000000", bold=True)

            # Apply styling to DOC columns efficiently
            doc_cols = [c for c in df.columns if "DOC" in c]
            if doc_cols:
                col_indices = [df.columns.get_loc(c) + 1 for c in doc_cols]
                
                # Pre-fetch colors to avoid dictionary overhead in loop
                c_critical = colors["critical"]
                c_low = colors["low"]
                c_optimal = colors["optimal"]
                c_monitor = colors["monitor"]
                c_high = colors["high"]
                c_excess = colors["excess"]
                c_overstock = colors["overstock"]

                for c_idx in col_indices:
                    for r in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=r, column=c_idx)
                        try:
                            val = float(cell.value)
                            if val <= 7:
                                cell.fill, cell.font = c_critical, font_white
                            elif val <= 15:
                                cell.fill, cell.font = c_low, font_black
                            elif val <= 30:
                                cell.fill, cell.font = c_optimal, font_white
                            elif val <= 45:
                                cell.fill, cell.font = c_monitor, font_black
                            elif val <= 60:
                                cell.fill, cell.font = c_high, font_black
                            elif val <= 90:
                                cell.fill, cell.font = c_excess, font_white
                            else:
                                cell.fill, cell.font = c_overstock, font_white
                        except:
                            pass
                        
            # Header styling
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
    return output.getvalue()

if process_btn:
    if not (pm_file and inv_file):
        st.error("Please upload Purchase Master and Inventory Report.")
    else:
        with st.spinner("Processing reports..."):
            # 1. Process Inventory
            inventory_df = pd.read_csv(inv_file)
            
            # Find ASIN column in Inventory
            asin_col_inv = next((c for c in ["asin", "ASIN"] if c in inventory_df.columns), None)
            if not asin_col_inv:
                st.error("Could not find an ASIN column in Inventory report. Expected 'asin' or 'ASIN'.")
                st.stop()
            
            # Normalize ASIN
            inventory_df[asin_col_inv] = inventory_df[asin_col_inv].astype(str).str.strip().str.upper()

            # Find SKU column in Inventory (for reference)
            sku_col_inv = next((c for c in ["sku", "Seller SKU", "SKU"] if c in inventory_df.columns), None)
            if sku_col_inv:
                inventory_df[sku_col_inv] = inventory_df[sku_col_inv].astype(str).str.strip().str.upper()

            # Process Reserved Report if provided
            reserved_df = pd.DataFrame()
            if res_file:
                reserved_df = pd.read_csv(res_file)
                asin_col_res = next((c for c in ["asin", "ASIN"] if c in reserved_df.columns), None)
                if not asin_col_res:
                    # Fallback to looking for SKU if ASIN is missing in reserved, but user said "from reserve in file column asin"
                    st.error("Could not find an ASIN column in Reserved Inventory report.")
                    st.stop()
                
                reserved_df[asin_col_res] = reserved_df[asin_col_res].astype(str).str.strip().str.upper()
                
                # Ensure numeric
                for col in ["reserved_customerorders", "reserved_fc-transfers", "reserved_fc-processing"]:
                    reserved_df[col] = clean_numeric_col(reserved_df, col)
                
                # Pivot Reserved by ASIN
                reserved_df = reserved_df.groupby(asin_col_res)[["reserved_customerorders", "reserved_fc-transfers", "reserved_fc-processing"]].sum().reset_index()
                reserved_df = reserved_df.rename(columns={asin_col_res: "(Parent) ASIN"})

            # Merge Reserved into Inventory (Pre-aggregation)
            # rename inventory asin to (Parent) ASIN for consistency
            inventory_df = inventory_df.rename(columns={asin_col_inv: "(Parent) ASIN"})
            
            # We don't join yet, we will pivot inventory by ASIN first to ensure match
            # But we need basic cleaning on inventory metrics
            inventory_df["afn-warehouse-qty"] = clean_numeric_col(inventory_df, "afn-warehouse-quantity")
            inventory_df["afn-reserved-qty"] = clean_numeric_col(inventory_df, "afn-reserved-quantity")
            inventory_df["Transfer Stock"] = (
                clean_numeric_col(inventory_df, "afn-inbound-working-quantity") + 
                clean_numeric_col(inventory_df, "afn-inbound-shipped-quantity") +
                clean_numeric_col(inventory_df, "afn-inbound-receiving-quantity")
            )

            # Pivot Inventory by ASIN
            inv_agg_dict = {
                "afn-warehouse-qty": "sum",
                "afn-reserved-qty": "sum",
                "Transfer Stock": "sum"
            }
            if sku_col_inv:
                inv_agg_dict[sku_col_inv] = "first" # Keep one SKU for reference
            
            inv_pivoted = inventory_df.groupby("(Parent) ASIN").agg(inv_agg_dict).reset_index()
            if sku_col_inv:
                inv_pivoted = inv_pivoted.rename(columns={sku_col_inv: "SKU"})

            # Now merge Reserved with Pivoted Inventory
            if not reserved_df.empty:
                inv_subset = inv_pivoted.merge(reserved_df, on="(Parent) ASIN", how="outer")
            else:
                inv_subset = inv_pivoted.copy()
                inv_subset["reserved_customerorders"] = 0
                inv_subset["reserved_fc-transfers"] = 0
                inv_subset["reserved_fc-processing"] = 0

            # Fill NaNs for numeric columns after merge
            num_fields = ["afn-warehouse-qty", "afn-reserved-qty", "Transfer Stock", 
                          "reserved_customerorders", "reserved_fc-transfers", "reserved_fc-processing"]
            for col in num_fields:
                if col in inv_subset.columns:
                    inv_subset[col] = inv_subset[col].fillna(0)
                else:
                    inv_subset[col] = 0

            # Calculate Total Stock at ASIN level
            inv_subset["Total Stock"] = (
                inv_subset["afn-warehouse-qty"] - 
                inv_subset["reserved_customerorders"] +
                inv_subset["reserved_fc-transfers"] + 
                inv_subset["reserved_fc-processing"]
            )

            # 2. Process Purchase Master
            pm_df = pd.read_excel(pm_file)
            
            # Find ASIN column in PM
            asin_col_pm = next((c for c in ["ASIN", "(Parent) ASIN", "asin"] if c in pm_df.columns), None)
            if not asin_col_pm:
                st.error("Could not find an ASIN column in Purchase Master.")
                st.stop()
            
            # Normalize ASIN
            pm_df[asin_col_pm] = pm_df[asin_col_pm].astype(str).str.strip().str.upper()
            pm_df = pm_df.rename(columns={asin_col_pm: "(Parent) ASIN"})

            # Find SKU column in PM
            possible_sku_cols = ["Seller SKU", "Amazon Sku Name", "SKU", "EasycomSKU", "sku", "Amazon Sku"]
            sku_col_pm = next((c for c in possible_sku_cols if c in pm_df.columns), None)
            if sku_col_pm:
                pm_df[sku_col_pm] = pm_df[sku_col_pm].astype(str).str.strip().str.upper()
                pm_df = pm_df.rename(columns={sku_col_pm: "SKU"})

            # Columns to keep (with safety checks)
            meta_cols = {
                "Brand": "Brand",
                "Product Name": "Product Name",
                "Brand Manager": "Brand Manager",
                "CP": "CP",
                "MRP": "MRP",
                "Vendor SKU Codes": "Vendor SKU Codes"
            }
            
            # Pivot PM by ASIN
            pm_agg_dict = {}
            for orig, target in meta_cols.items():
                if orig in pm_df.columns:
                    if target in ["CP", "MRP"]:
                        # Pre-clean numeric
                        pm_df[orig] = pd.to_numeric(
                            pm_df[orig].astype(str).str.replace(",", "", regex=False).str.strip(), 
                            errors="coerce"
                        ).fillna(0)
                        pm_agg_dict[orig] = "mean" # Cost Price/MRP should be averaged or first
                    else:
                        pm_agg_dict[orig] = "first"
            
            if "SKU" in pm_df.columns:
                pm_agg_dict["SKU"] = "first" # Keep one SKU for reference

            pm_subset = pm_df.groupby("(Parent) ASIN").agg(pm_agg_dict).reset_index()
            pm_subset = pm_subset.rename(columns={k:v for k,v in meta_cols.items() if k in pm_subset.columns})

            # 3. Process Business Reports (Starting consolidation)
            # Merge PM and Inventory on (Parent) ASIN
            consolidated_df = pm_subset.merge(inv_subset, on="(Parent) ASIN", how="outer", suffixes=('', '_inv'))
            
            # Handle SKU consolidation
            if "SKU_inv" in consolidated_df.columns:
                if "SKU" not in consolidated_df.columns:
                    consolidated_df["SKU"] = consolidated_df["SKU_inv"]
                else:
                    consolidated_df["SKU"] = consolidated_df["SKU"].fillna(consolidated_df["SKU_inv"])
                consolidated_df.drop(columns=["SKU_inv"], inplace=True)
            
            consolidated_df["Total Stock"] = consolidated_df["Total Stock"].fillna(0)
            consolidated_df["Transfer Stock"] = consolidated_df["Transfer Stock"].fillna(0)
            
            drr_cols = []
            
            for days in intervals:
                if br_files[days]:
                    br_pivot = process_br(br_files[days], days)
                    # br_pivot is already keyed by (Parent) ASIN
                    consolidated_df = consolidated_df.merge(br_pivot, on="(Parent) ASIN", how="outer")
                    
                    # Consolidated sales/views NaN filling
                    consolidated_df[f"{days} Days Sales Qty"] = consolidated_df[f"{days} Days Sales Qty"].fillna(0)
                    consolidated_df[f"{days} Days Page Views"] = consolidated_df[f"{days} Days Page Views"].fillna(0)
                    
                    # Vectorized calculations for speed
                    drr_val = (consolidated_df[f"{days} Days Sales Qty"] / days).round(2)
                    consolidated_df[f"{days} Days DRR"] = drr_val
                    
                    # Vectorized DOC calculation (fast)
                    consolidated_df[f"{days} Days DOC"] = (
                        (consolidated_df["Total Stock"] / drr_val)
                        .replace([np.inf, -np.inf], 0)
                        .fillna(0)
                        .round(0)
                        .astype(int)
                    )
                    drr_cols.append(f"{days} Days DRR")

            # 4. Final Data Enrichment (Listing Status & seller-sku)
            consolidated_df["Listing Status"] = ""
            consolidated_df["seller-sku"] = ""
            if list_file:
                if list_file.name.endswith(".csv"):
                    listing_df = pd.read_csv(list_file)
                else:
                    listing_df = pd.read_excel(list_file)
                
                # Use 4th column for SKU as per cl.py logic
                if len(listing_df.columns) >= 4:
                    sku_series = listing_df.iloc[:, 3].astype(str).str.strip().str.upper()
                    listing_skus_dict = dict(zip(sku_series, sku_series))
                    
                    # Apply to all SKUs in consolidated_df
                    temp_sku = consolidated_df["SKU"].astype(str).str.strip().str.upper()
                    consolidated_df["seller-sku"] = temp_sku.map(listing_skus_dict).fillna("")
                    consolidated_df["Listing Status"] = temp_sku.apply(lambda x: "Closed" if x in listing_skus_dict else "")

            # Final safety check for all numeric columns (once, outside loop)
            num_cols_final = ["Total Stock", "Transfer Stock", "afn-warehouse-qty", "afn-reserved-qty", 
                              "reserved_customerorders", "reserved_fc-transfers", "reserved_fc-processing", "CP", "MRP"]
            for col in num_cols_final:
                if col in consolidated_df.columns:
                    consolidated_df[col] = consolidated_df[col].fillna(0)

            # 5. Aggregate Metrics and Specialized Reports
            if drr_cols:
                consolidated_df["Max DRR"] = consolidated_df[drr_cols].max(axis=1)
                consolidated_df["Avg DRR"] = consolidated_df[drr_cols].mean(axis=1).round(2)
                
                # Vectorized DOC (Max)
                consolidated_df["DOC (Max)"] = (
                    (consolidated_df["Total Stock"] / consolidated_df["Max DRR"])
                    .replace([np.inf, -np.inf], 0)
                    .fillna(0)
                    .round(0)
                    .astype(int)
                )

            # Calculated Financial Columns
            if "CP" in consolidated_df.columns:
                consolidated_df["CP As Per Total Stock Qty"] = (consolidated_df["CP"] * consolidated_df["Total Stock"]).round(2)
                for days in intervals:
                    if f"{days} Days Sales Qty" in consolidated_df.columns:
                        consolidated_df[f"{days} CP As Per Total Sale Qty"] = (consolidated_df["CP"] * consolidated_df[f"{days} Days Sales Qty"]).round(2)

            # Define Layout order precisely as requested by the user
            # SKU, (Parent) ASIN, Vendor SKU Codes, Brand, Brand Manager, Product Name, CP, 
            # afn-warehouse-qty, afn-reserved-qty, reserved_customerorders, reserved_fc-transfers, reserved_fc-processing, 
            # Total Stock, CP As Per Total Stock Qty, ... multi-day ...
            # seller-sku, Listing Status
            layout_meta = ["(Parent) ASIN", "SKU", "Vendor SKU Codes", "Brand", "Brand Manager", "Product Name", "CP"]
            
            layout_stock = ["afn-warehouse-qty", "afn-reserved-qty", "reserved_customerorders", 
                           "reserved_fc-transfers", "reserved_fc-processing", "Total Stock", "CP As Per Total Stock Qty"]
            
            layout_metrics = []
            for days in intervals:
                # {day} Days Sales Qty, {day} CP As Per Total Sale Qty, {day} Days DRR, {day} Days DOC, {day} Days Page Views
                prefix = f"{days} Days"
                cp_prefix = f"{days} CP As Per Total Sale Qty"
                if f"{prefix} Sales Qty" in consolidated_df.columns:
                    layout_metrics.extend([
                        f"{prefix} Sales Qty",
                        cp_prefix,
                        f"{prefix} DRR", 
                        f"{prefix} DOC", 
                        f"{prefix} Page Views"
                    ])
            
            ordered_cols = [c for c in (layout_meta + layout_stock + layout_metrics + ["seller-sku", "Listing Status"]) if c in consolidated_df.columns]
            final_df = consolidated_df[ordered_cols].copy()
            
            # Create Specialized Reports (All tabs use the uniform layout `final_df`)
            inv_view = final_df.copy()
            
            # OOS View (Fulfillable == 0 or Total Stock == 0)
            oos_view = final_df[final_df["Total Stock"] == 0].copy()
            
            # Overstock View (Any interval DOC > 90 or DOC (Max) > 90)
            # Find all DOC columns
            doc_cols = [c for c in final_df.columns if "DOC" in c]
            if doc_cols:
                overstock_view = final_df[final_df[doc_cols].gt(90).any(axis=1)].copy()
            else:
                overstock_view = final_df[final_df["Total Stock"] > 500].copy() # Fallback
            
            # Create Pivot Tables for OOS and Overstock
            oos_pivot = create_stock_pivot(oos_view)
            overstock_pivot = create_stock_pivot(overstock_view)
            
            st.success("Analysis Complete!")
            
            tab1, tab2, tab3, tab4 = st.tabs(["📊 Consolidated View", "📦 Inventory View", "⚠️ OOS Analysis", "📈 Overstock"])
            
            with tab1:
                st.subheader("Multi-Day Consolidated Analysis")
                st.dataframe(apply_doc_styling(final_df))
            
            with tab2:
                st.subheader("Inventory Breakdown (Full Layout)")
                st.dataframe(apply_doc_styling(inv_view))
                
            with tab3:
                st.subheader("Out of Stock (OOS) Items")
                st.dataframe(apply_doc_styling(oos_view))
                st.divider()
                st.subheader("OOS Pivot Table")
                st.dataframe(apply_doc_styling(oos_pivot))
                
            with tab4:
                st.subheader("Overstock Analysis (DOC > 90)")
                st.dataframe(apply_doc_styling(overstock_view))
                st.divider()
                st.subheader("Overstock Pivot Table")
                st.dataframe(apply_doc_styling(overstock_pivot))
            
            # Excel Download
            excel_sheets = {
                "Consolidated": final_df,
                "Inventory": inv_view,
                "OOS_Report": oos_view,
                "OOS_Pivot": oos_pivot,
                "Overstock_Report": overstock_view,
                "Overstock_Pivot": overstock_pivot
            }
            excel_data = create_excel(excel_sheets, intervals)
            st.download_button(
                label="📥 Download All Reports (Excel)",
                data=excel_data,
                file_name=f"Multi_Day_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
